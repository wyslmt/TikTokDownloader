import csv
import logging
import os
import sqlite3
import time

from openpyxl import Workbook
from openpyxl import load_workbook

from src.StringCleaner import Cleaner


class BaseLogger:
    """不记录日志，空白日志记录器"""

    def __init__(self):
        self.log = None  # 记录器主体
        self._root = "./"  # 日志记录保存根路径
        self._folder = "Log"  # 日志记录保存文件夹名称
        self._name = "%Y-%m-%d %H.%M.%S"  # 日志文件名称

    @property
    def root(self):
        return self._root

    @root.setter
    def root(self, value):
        pass

    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, value):
        pass

    @property
    def folder(self):
        return self._folder

    @folder.setter
    def folder(self, value: str):
        pass

    def run(self, *args, **kwargs):
        pass

    @staticmethod
    def info(text: str, output=True):
        if output:
            print(text)

    @staticmethod
    def warning(text: str, output=True):
        if output:
            print(text)

    @staticmethod
    def error(text: str, output=True):
        if output:
            print(text)


class LoggerManager(BaseLogger):
    """日志记录"""

    def __init__(self):
        super().__init__()

    @property
    def root(self):
        return self._root

    @root.setter
    def root(self, value):
        if os.path.exists(value) and os.path.isdir(value):
            self._root = value
        else:
            print("日志保存路径错误！将使用当前路径作为日志保存路径！")
            self._root = "./"

    @property
    def name(self):
        return self._name

    @name.setter
    def name(self, value):
        if value:
            try:
                _ = time.strftime(value, time.localtime())
                self._name = value
            except ValueError:
                print("日志名称格式错误，将使用默认时间格式（年-月-日 时.分.秒）")
                self._name = "%Y-%m-%d %H.%M.%S"
        else:
            print("日志名称格式错误，将使用默认时间格式（年-月-日 时.分.秒）")
            self._name = "%Y-%m-%d %H.%M.%S"

    @property
    def folder(self):
        return self._folder

    @folder.setter
    def folder(self, value: str):
        self._folder = s if (s := Cleaner().filter(value)) else "Log"

    def run(
            self,
            format_="%(asctime)s[%(levelname)s]:  %(message)s", filename=None):
        if not os.path.exists(dir_ := os.path.join(self.root, self.folder)):
            os.mkdir(dir_)
        self.log = logging
        self.log.basicConfig(
            filename=os.path.join(
                dir_,
                filename or f"{time.strftime(self.name, time.localtime())}.log"),
            level=logging.INFO,
            datefmt='%Y-%m-%d %H:%M:%S',
            format=format_,
            encoding="UTF-8")

    def info(self, text: str, output=True):
        if output:
            print(text)
        self.log.info(text)

    def warning(self, text: str, output=True):
        if output:
            print(text)
        self.log.warning(text)

    def error(self, text: str, output=True):
        if output:
            print(text)
        self.log.error(text)


class NoneLogger:
    def __init__(self, *args, **kwargs):
        pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        pass

    def save(self, *args, **kwargs):
        pass


class CSVLogger:
    """CSV格式记录"""

    def __init__(
            self,
            root: str,
            title_line,
            name="Download",
            *args,
            **kwargs):
        self.file = None  # 文件对象
        self.writer = None  # CSV对象
        self.root = root  # 文件路径
        self.name = name  # 文件名称
        self.title_line = title_line  # 标题行

    def __enter__(self):
        if not os.path.exists(self.root):
            os.mkdir(self.root)
        self.root = os.path.join(self.root, f"{self.name}.csv")
        self.file = open(self.root,
                         "a",
                         encoding="UTF-8",
                         newline="")
        self.writer = csv.writer(self.file)
        self.title()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.file.close()

    def title(self):
        if os.path.getsize(self.root) == 0:
            # 如果文件没有任何数据，则写入标题行
            self.save(self.title_line)

    def save(self, data):
        self.writer.writerow(data)


class XLSXLogger:
    """XLSX格式"""

    def __init__(
            self,
            root: str,
            title_line,
            name="Download",
            *args,
            **kwargs):
        self.book = None  # XLSX数据簿
        self.sheet = None  # XLSX数据表
        self.root = root  # 文件路径
        self.name = name  # 文件名称
        self.title_line = title_line  # 标题行

    def __enter__(self):
        if not os.path.exists(self.root):
            os.mkdir(self.root)
        self.root = os.path.join(self.root, f"{self.name}.xlsx")
        if os.path.exists(self.root):
            self.book = load_workbook(self.root)
        else:
            self.book = Workbook()
        self.sheet = self.book.active
        self.title()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.book.save(self.root)
        self.book.close()

    def title(self):
        if not self.sheet["A1"].value:
            # 如果文件没有任何数据，则写入标题行
            for col, value in enumerate(self.title_line, start=1):
                self.sheet.cell(row=1, column=col, value=value)

    def save(self, data):
        self.sheet.append(data)


class SQLLogger:
    """SQLite保存数据"""

    def __init__(
            self,
            root: str,
            file,
            title_line,
            title_type,
            name="Download", ):
        self.db = None  # 数据库
        self.cursor = None  # 游标对象
        self.root = root  # 文件路径
        self.name = name  # 数据表名称
        self.file = file  # 数据库文件名称
        self.title_line = title_line  # 数据表列名
        self.title_type = title_type  # 数据表数据类型

    def __enter__(self):
        if not os.path.exists(self.root):
            os.mkdir(self.root)
        self.db = sqlite3.connect(
            os.path.join(
                self.root, self.file
            ))
        self.cursor = self.db.cursor()
        self.create()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.db.close()

    def create(self):
        create_sql = f"""CREATE TABLE IF NOT EXISTS {self.name} ({", ".join([f"{i} {j}" for i, j in zip(self.title_line, self.title_type)])});"""
        self.cursor.execute(create_sql)
        self.db.commit()

    def save(self, data):
        insert_sql = f"""REPLACE INTO {self.name} ({", ".join(self.title_line)}) VALUES ({", ".join(["?" for _ in self.title_line])});"""
        self.cursor.execute(insert_sql, data)
        self.db.commit()


class RecordManager:
    """检查数据储存路径和文件夹"""
    clean = Cleaner()
    Title = (
        "作品类型",
        "采集时间",
        "作品ID",
        "作品描述",
        "发布时间",
        "账号昵称",
        "Video_ID",
        "点赞数量",
        "评论数量",
        "收藏数量",
        "分享数量")
    Type_ = (
        "CHARACTER(2) NOT NULL",
        "CHARACTER(20) NOT NULL",
        "CHARACTER(19) PRIMARY KEY",
        "CHARACTER(128) NOT NULL",
        "CHARACTER(20) NOT NULL",
        "CHARACTER(20) NOT NULL",
        "CHARACTER(64)",
        "INTEGER NOT NULL",
        "INTEGER NOT NULL",
        "INTEGER NOT NULL",
        "INTEGER NOT NULL",
    )
    Comment_Title = (
        "采集时间",
        "评论ID",
        "评论时间",
        "用户昵称",
        "IP归属地",
        "评论内容",
        "评论图片",
        "点赞数量",
        "回复数量",
        "回复ID",
    )
    Comment_Type = (
        "CHARACTER(20) NOT NULL",
        "CHARACTER(19) PRIMARY KEY",
        "CHARACTER(20) NOT NULL",
        "CHARACTER(20) NOT NULL",
        "CHARACTER(10) NOT NULL",
        "CHARACTER(256) NOT NULL",
        "CHARACTER(256) NOT NULL",
        "INTEGER NOT NULL",
        "INTEGER NOT NULL",
        "CHARACTER(19) NOT NULL",
    )
    User_Title = (
        "ID",
        "采集时间",
        "昵称",
        "简介",
        "抖音号",
        "年龄",
        "标签",
        "企业",
        "sec_uid",
        "uid",
        "short_id",
        "头像",
        "背景图",
        "作品数量",
        "获赞数量",
        "喜欢作品数量",
        "粉丝数量",
        "关注数量",
        "粉丝最多数量",
    )
    User_Type = (
        "INTEGER PRIMARY KEY",
        "CHARACTER(20) NOT NULL",
        "CHARACTER(20) NOT NULL",
        "TEXT",
        "TEXT",
        "INTEGER",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "TEXT",
        "CHARACTER(256) NOT NULL",
        "CHARACTER(256) NOT NULL",
        "INTEGER",
        "INTEGER",
        "INTEGER",
        "INTEGER",
        "INTEGER",
        "INTEGER",
    )
    DataSheet = {
        "comment": {
            "file": "CommentData.db",
            "title_line": Comment_Title,
            "title_type": Comment_Type},
        "user": {
            "file": "UserData.db",
            "title_line": User_Title,
            "title_type": User_Type},
        "": {
            "file": "TikTokDownloader.db",
            "title_line": Title,
            "title_type": Type_
        },
        "mix": {
            "file": "MixData.db",
            "title_line": Title,
            "title_type": Type_
        }
    }
    DataLogger = {
        "csv": CSVLogger,
        "xlsx": XLSXLogger,
        "sql": SQLLogger,
        "": NoneLogger,
    }

    def run(self, root="./", folder="Data", type_="", format_=""):
        root = root if os.path.exists(root) else "./"
        name = os.path.join(root, self.clean.filter(folder) or "Data")
        type_ = self.DataSheet.get(type_, {})
        format_ = self.DataLogger.get(format_)
        return format_, name, type_
